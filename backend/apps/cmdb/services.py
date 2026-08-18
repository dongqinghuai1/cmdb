"""cmdb services: attrs validation / rack placement / excel import-export (PRD 5.5)."""
import io
from datetime import datetime

from apps.cmdb.models import CiModelAttr, Device


class DeviceService:
    BUILTIN = {"sn", "asset_no", "name", "hostname", "vendor", "hw_model", "sw_version",
               "manage_ip", "rack_start_u", "rack_units", "rated_power_w", "remark"}

    @classmethod
    def validate_attrs(cls, ci_model, attrs: dict):
        errors = {}
        for a in ci_model.attrs_def.all():
            v = attrs.get(a.code, None)
            if v in (None, ""):
                if a.is_required and not a.default_value:
                    errors[a.code] = "required"
                continue
            try:
                if a.attr_type == CiModelAttr.AttrType.INT:
                    int(v)
                elif a.attr_type == CiModelAttr.AttrType.FLOAT:
                    float(v)
                elif a.attr_type == CiModelAttr.AttrType.BOOL:
                    assert str(v).lower() in ("true", "false", "1", "0")
                elif a.attr_type == CiModelAttr.AttrType.ENUM:
                    assert str(v) in [str(x) for x in a.enum_options], f"must be one of {a.enum_options}"
                elif a.attr_type == CiModelAttr.AttrType.DATE:
                    datetime.strptime(str(v)[:10], "%Y-%m-%d")
            except Exception as e:
                errors[a.code] = f"type {a.attr_type} check failed: {e}"
        for k in attrs:
            if k in cls.BUILTIN:
                errors[k] = f"conflicts with builtin field: {k}"
        if errors:
            raise ValueError("; ".join(f"{k}:{v}" for k, v in errors.items()))
        return attrs

    @staticmethod
    def place(device: Device, rack, start_u: int):
        from apps.dcim.services import RackService
        units = device.rack_units or 1
        RackService.check_placement(rack.id, start_u, units, exclude_device_id=device.id)
        device.rack = rack
        device.rack_start_u = start_u
        device.save(update_fields=["rack", "rack_start_u", "region", "updated_at"])

    IMPORT_COLUMNS = ["name", "model_code", "vendor", "sn", "asset_no", "manage_ip",
                      "site_code", "rack_name", "start_u", "units", "hw_model", "owner"]

    @classmethod
    def import_excel(cls, fh, user) -> dict:
        from openpyxl import load_workbook
        from apps.dcim.models import Rack, Site
        from apps.cmdb.models import CiModel
        wb = load_workbook(fh, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        result = {"total": len(rows), "success": 0, "failed": 0, "errors": []}
        for i, row in enumerate(rows, start=2):
            data = dict(zip(cls.IMPORT_COLUMNS, [str(c).strip() if c is not None else "" for c in row]))
            try:
                ci = CiModel.objects.get(code=data["model_code"])
                site = Site.objects.get(code=data["site_code"]) if data["site_code"] else None
                if not data["name"]:
                    raise ValueError("name is required")
                dev = Device(
                    name=data["name"], model=ci, vendor=data["vendor"],
                    sn=data["sn"] or None, asset_no=data["asset_no"] or None,
                    manage_ip=data["manage_ip"] or None, hw_model=data["hw_model"],
                    site=site, region_id=site.region_id if site else None,
                    rack_start_u=int(data["start_u"]) if data["start_u"] else None,
                    rack_units=int(data["units"]) if data["units"] else ci.default_u_height,
                )
                if site and data["rack_name"]:
                    dev.rack = Rack.objects.get(site=site, name=data["rack_name"])
                    cls.place(dev, dev.rack, dev.rack_start_u)
                else:
                    dev.save()
                result["success"] += 1
            except Exception as e:
                result["failed"] += 1
                result["errors"].append({"row": i, "error": str(e)[:200]})
        return result

    @classmethod
    def export_excel(cls, qs) -> io.BytesIO:
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active; ws.append(cls.IMPORT_COLUMNS)
        for d in qs.select_related("model", "site", "rack", "owner"):
            ws.append([d.name, d.model.code, d.vendor, d.sn, d.asset_no, d.manage_ip,
                       d.site.code if d.site else "", d.rack.name if d.rack else "",
                       d.rack_start_u or "", d.rack_units, d.hw_model,
                       d.owner.username if d.owner_id else ""])
        buf = io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf
